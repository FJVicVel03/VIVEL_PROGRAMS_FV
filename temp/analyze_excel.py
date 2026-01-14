import openpyxl

# Analyze PROCTOR sheet from ZANJA file
file_path = r'c:\Users\ferna\Desktop\RepositorioLocal\VIVEL_PROGRAMS_FV\PROCTOR Y DENSIDADES ZANJA.xlsx'

print(f"FILE: {file_path.split('\\')[-1]}")
print('='*60)

wb = openpyxl.load_workbook(file_path)
print(f"Sheets: {wb.sheetnames}")

if 'PROCTOR' in wb.sheetnames:
    ws = wb['PROCTOR']
    print(f"\n--- PROCTOR Sheet ---")
    print(f"Dimensions: {ws.dimensions}")
    
    # Print all cells with content (rows 1-50)
    print("\nCells with content:")
    for row_idx in range(1, 50):
        row_content = []
        for col_idx in range(1, 45):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                val = str(cell.value)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_content.append(f"{cell.coordinate}=[F]{val[:50]}")
                else:
                    row_content.append(f"{cell.coordinate}={val[:25]}")
        if row_content:
            print(f"  Row {row_idx}: {', '.join(row_content)}")
